# limit the number of cpus used by high performance libraries
import os
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["VECLIB_MAXIMUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"

import sys
# sys.path.insert(0, './yolov5')
sys.path.append('./Detection/yolov5')
sys.path.append('./Detection/deep/reid')

import argparse
import numpy as np
import cv2
import torch
from Track.tracker import Tracker, get_all_detections
import openpyxl

from dataset.self_dataset2 import self_dataset2
from Detection.detection import YOLO_reid_model
from TiProcess.radar_process import get_dbscan_points, get_radar_features
from Tools.LIdarFilter import lidar2img_filter
import matplotlib.pyplot as plt

from TiProcess.BEV_TI import pts2rbev, OCUpts2rbev


def draw_detections_in_BEV(detections_r, detections_c, lidar_json, OCU_json, rgb_json, curr_label, group_num):
    """
        Lidar_annotation_points: (N_gt_tracks, 5) id, xyz,classes  x:left-right, y:front-behind
        radar_center: (N_radar_dets, 2) xy  x:left-right, y:front-behind
        cam_proj_center: (N_img_dets, 2) xy  x:left-right, y:front-behind
    """
    Lidar_annotation_points = OCUpts2rbev(lidar_json, OCU_json, rgb_json, curr_label)
    radar_center = np.array([detections_r[i].center for i in range(len(detections_r))])
    cam_proj_center = np.array([detections_c[i].center for i in range(len(detections_c))])

    # print('Lidar_annotation_points:', Lidar_annotation_points)
    # print('radar_center:', radar_center)
    # print('cam_proj_center:', cam_proj_center, '\n')

    plt.xlim(-20, 20)
    plt.ylim(0, 40)
    for i in range(Lidar_annotation_points.shape[0]):
        x = Lidar_annotation_points[i][1]
        y = Lidar_annotation_points[i][2]
        plt.scatter(x, y, s=50, marker='x', c='k')

    for i in range(radar_center.shape[0]):
        x = radar_center[i][0]
        y = radar_center[i][1]
        plt.scatter(x, y, s=50, marker='^', c='b')

    for i in range(cam_proj_center.shape[0]):
        x = cam_proj_center[i][0]
        y = cam_proj_center[i][1]
        plt.scatter(x, y, s=50, marker='s', c='r')

    plt.title(group_num)
    plt.show()
    plt.pause(0.1)
    # save_name = '/media/personal_data/zhangq/RadarRGBFusionNet/results/' + folder + '.png'
    # plt.savefig(save_name, dpi=500)
    plt.clf()


def get_detections_and_labels(detections_r, detections_c, lidar_json, OCU_json, rgb_json, curr_label, Lidar_points, frame_num):
    """
            Lidar_annotation_points: (N_gt_tracks, 5) xyz,id,classes  x:left-right, y:front-behind
            radar_center: (N_radar_dets, 2) xy  x:left-right, y:front-behind
            cam_proj_center: (N_img_dets, 2) xy  x:left-right, y:front-behind
            Lidar_points: (N_Lidar_points_num, 5), x y z intensity idx_laser
        """
    image_width = rgb_json['image_size'][1]
    image_height = rgb_json['image_size'][0]
    Lidar_annotation_points = lidar2img_filter(lidar_json, OCU_json, rgb_json, curr_label, image_width, image_height, Lidar_points)
    # Lidar_annotation_points2 = OCUpts2rbev(lidar_json, OCU_json, rgb_json, curr_label)
    radar_center = np.array([detections_r[i].center for i in range(len(detections_r))])
    cam_proj_center = np.array([detections_c[i].center for i in range(len(detections_c))])

    # add frame_num
    # print('lens:', Lidar_annotation_points.shape[0], radar_center.shape[0], cam_proj_center.shape[0])
    len_lidar = Lidar_annotation_points.shape[0]
    len_radar = radar_center.shape[0]
    len_img = cam_proj_center.shape[0]
    if len_lidar != 0:
        Lidar_annotations = np.hstack(
            (Lidar_annotation_points, np.ones((len_lidar, 1)) * frame_num, np.ones((len_lidar, 1)) * -1))
    else:
        Lidar_annotations = np.zeros(shape=(0, 7))

    if len_radar != 0:
        radar_dets = np.hstack((radar_center, np.ones((len_radar, 1)) * frame_num, np.ones((len_radar, 1)) * -1))
    else:
        radar_dets = np.zeros(shape=(0, 4))

    if len_img != 0:
        img_dets = np.hstack((cam_proj_center, np.ones((len_img, 1)) * frame_num, np.ones((len_img, 1)) * -1))
    else:
        img_dets = np.zeros(shape=(0, 4))

    return Lidar_annotations, radar_dets, img_dets

def run(opt):
    out, source, yolo_model, reid_model, show_vid, save_vid, save_txt, imgsz, evaluate, half, project, name, exist_ok, \
    device, classes, conf_thres, iou_thres, max_det = opt.output, opt.source, opt.yolo_model, opt.deep_sort_model,\
     opt.show_vid, opt.save_vid, opt.save_txt, opt.imgsz, opt.evaluate, opt.half, opt.project, opt.name, opt.exist_ok,\
     opt.device, opt.classes, opt.conf_thres, opt.iou_thres, opt.max_det

    # initialize parameter

    # Tracker
    tracker = Tracker()

    # detection and reid model
    bs = 1  # batch_size
    detection_model = YOLO_reid_model(yolo_model, reid_model, device=device, imgsz=imgsz, classes=classes,
                                      conf_thres=conf_thres, iou_thres=iou_thres, max_det=max_det)

    # Dataloader

    file_path = '/mnt/ChillDisk/personal_data/zhangq/RadarRGBFusionNet2_20231128/GroupPath/Test.xlsx'
    DataPath = openpyxl.load_workbook(file_path)
    ws = DataPath.active
    groups_excel = ws['A']
    label_excel = ws['B']
    datasets_path = []
    for cell in groups_excel:
        datasets_path.append(cell.value)

    group_num = -1
    all_estimated_tracks = np.zeros(shape=(0, 6))  # save tracking result, (id,x,y,o, frame_num, group_num)
    all_gt_tracks = np.zeros(shape=(0, 7))  # save gts, (id,x,y,z, classes, frame_num, group_num)
    all_radar_detections = np.zeros(shape=(0, 4))  # save radar dets, (x,y, frame_num, group_num)
    all_img_detections = np.zeros(shape=(0, 4))  # save img dets, (x,y, frame_num, group_num)
    group_names = np.zeros(shape=(0, 2))     # (group_num, group_name)
    # sum_2D_GT = 0
    # sum_3D_GT = 0
    for i in range(len(datasets_path)):
        # data_base_path = '/media/personal_data/zhangq/RadarRGBFusionNet2_20231128/dataset/datas'
        # label_base_path = '/media/personal_data/zhangq/RadarRGBFusionNet2_20231128/dataset/labels'

        data_base_path = '/mnt/ourDataset_v2/ourDataset_v2_label'
        source = os.path.join(data_base_path, datasets_path[i])

        dataset = self_dataset2(source, img_size=imgsz)
        group_num += 1
        print('group:', group_num, ', ', source)

        # if group_num < 4:
        #     continue
        # initialize parameter
        frame_num = -1
        estimated_tracks = np.zeros(shape=(0, 6))  # save tracking result, (id,x,y,o, frame_num, group_num)
        gt_tracks = np.zeros(shape=(0, 7))  # save gts, (id,x,y,z, classes, frame_num, group_num)
        radar_detections = np.zeros(shape=(0, 4))  # save radar dets, (x,y, frame_num, group_num)
        img_detections = np.zeros(shape=(0, 4))  # save img dets, (x,y, frame_num, group_num)
        group_name = [group_num, datasets_path[i]]
        group_names = np.vstack((group_names, group_name))

        for frame_idx, (rgb_img0, img, rgb_json, OCU_data, OCU_json, lidar_json, curr_label, Lidar_points) in enumerate(dataset):
            # print('frame_idx:', frame_idx)

            # if frame_idx == 138:
            #     print()
            outputs, detections_c, im0 = detection_model.get_detections(rgb_img0, img, rgb_json, OCU_json, OCU_data)
            # detections_c = detection_model.get_GT_detections(rgb_img0, rgb_json, TI_json)

            # if TI_data.shape[0] != 0:
            #     TI_points, TI_db, TI_num_dbscan = get_dbscan_points(TI_data, 0.9, 7)
            #     detections_r = get_radar_features(TI_points, TI_db, TI_num_dbscan)
            #     all_detections = get_all_detections(detections_r, detections_c)
            # else:
            #     all_detections = detections_c
            TI_points, TI_db, TI_num_dbscan = get_dbscan_points(OCU_data, 0.8, 10)
            detections_r = get_radar_features(TI_points, TI_db, TI_num_dbscan)
            all_detections = get_all_detections(detections_r, detections_c)

            # print('all_detections', len(all_detections))

            # draw_detections_in_BEV(detections_r, detections_c, lidar_json, OCU_json, rgb_json, curr_label, group_num)
            Lidar_annotations, radar_dets, img_dets = get_detections_and_labels(detections_r, detections_c, lidar_json,
                                                                                OCU_json, rgb_json, curr_label, Lidar_points, frame_idx)
            # sum_2D_GT += len(detections_c)
            # sum_3D_GT += len(Lidar_annotations)
            # print('2D Box Num:', len(detections_c), '3D GT Num:', len(Lidar_annotations), '\n')
            # print('Lidar_annotations', Lidar_annotations)
            gt_tracks = np.vstack((gt_tracks, Lidar_annotations))
            radar_detections = np.vstack((radar_detections, radar_dets))
            img_detections = np.vstack((img_detections, img_dets))

            tracker.predict()
            tracks_save = tracker.update(all_detections)
            # tracker.collision_predict()
            tracks_save[:, 4] = frame_idx
            estimated_tracks = np.vstack((estimated_tracks, tracks_save))
            # print('tracks_save:\n', tracks_save)

            # cv2.imshow('test', im0)
            # cv2.waitKey(100)
            # if frame_idx == 138:
            #     # cv2.imshow('test', im0)
            #     # if cv2.waitKey(0) & 0xFF == 27:
            #     #     break
            #     detections_indices = np.arange(len(detections_c))
            #     all_measurements = np.asarray([detections_c[i].center[0:2] for i in detections_indices])
            #     print('*****************'*5)

        estimated_tracks[:, 5] = group_num
        gt_tracks[:, 6] = group_num
        radar_detections[:, 3] = group_num
        img_detections[:, 3] = group_num
        print('img_detections_num:', len(img_detections))


        all_estimated_tracks = np.vstack((all_estimated_tracks, estimated_tracks))
        all_gt_tracks = np.vstack((all_gt_tracks, gt_tracks))
        all_radar_detections = np.vstack((all_radar_detections, radar_detections))
        all_img_detections = np.vstack((all_img_detections, img_detections))

        print('group:', group_num, ', done.')

    # print('sum_2D_GT:', sum_2D_GT)
    # print('sum_3D_GT:', sum_3D_GT)
    # print('all_estimated_tracks:\n', all_estimated_tracks, '\n')
    # print('all_gt_tracks:\n', all_gt_tracks, '\n')
    # print('all_radar_detections:\n', all_radar_detections, '\n')
    # print('all_img_detections:\n', all_img_detections, '\n')

    np.save("./TrackResults/all_estimated_tracks.npy", all_estimated_tracks)
    np.save("./TrackResults/all_gt_tracks.npy", all_gt_tracks)
    np.save("./TrackResults/all_radar_detections.npy", all_radar_detections)
    np.save("./TrackResults/all_img_detections.npy", all_img_detections)
    np.save("./TrackResults/group_names.npy", group_names)




if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--yolo_model', nargs='+', type=str, default='yolov5m.pt', help='model.pt path(s)')
    parser.add_argument('--deep_sort_model', type=str, default='osnet_x1_0')   # osnet_x0_25
    parser.add_argument('--source', type=str, default='0', help='source')  # file/folder, 0 for webcam
    parser.add_argument('--output', type=str, default='inference/output', help='output folder')  # output folder
    parser.add_argument('--imgsz', '--img', '--img-size', nargs='+', type=int, default=[640], help='inference size h,w')
    parser.add_argument('--conf-thres', type=float, default=0.3, help='object confidence threshold')
    parser.add_argument('--iou-thres', type=float, default=0.5, help='IOU threshold for NMS')
    parser.add_argument('--fourcc', type=str, default='mp4v', help='output video codec (verify ffmpeg support)')
    parser.add_argument('--device', default='', help='cuda device, i.e. 0 or 0,1,2,3 or cpu')
    parser.add_argument('--show-vid', default=True, action='store_true', help='display tracking video results')
    parser.add_argument('--save-vid', action='store_true', help='save video tracking results')
    parser.add_argument('--save-txt', action='store_true', help='save MOT compliant results to *.txt')
    # class 0 is person, 1 is bycicle, 2 is car... 79 is oven
    parser.add_argument('--classes', nargs='+', type=int, help='filter by class: --class 0, or --class 16 17')
    parser.add_argument('--agnostic-nms', action='store_true', help='class-agnostic NMS')
    parser.add_argument('--augment', action='store_true', help='augmented inference')
    parser.add_argument('--evaluate', action='store_true', help='augmented inference')
    parser.add_argument("--config_deepsort", type=str, default="deep_sort/configs/deep_sort.yaml")
    parser.add_argument("--half", action="store_true", help="use FP16 half-precision inference")
    parser.add_argument('--visualize', action='store_true', help='visualize features')
    parser.add_argument('--max-det', type=int, default=1000, help='maximum detection per image')
    parser.add_argument('--dnn', action='store_true', help='use OpenCV DNN for ONNX inference')
    parser.add_argument('--project', default='/mnt/ChillDisk/personal_data/zhangq/RadarRGBFusionNet2_20231128/runs/track', help='save results to project/name')
    parser.add_argument('--name', default='exp', help='save results to project/name')
    parser.add_argument('--exist-ok', action='store_true', help='existing project/name ok, do not increment')
    opt = parser.parse_args()
    # opt.source = 'video/twopeople.mp4'
    # opt.source = 'video/206.mp4'
    # opt.source = '/media/ourDataset/v1.0_label/20211025_1_group0012_185frames_37labeled'
    # opt.source = '/media/ourDataset/v1.0_label/20211027_1_group0010_148frames_30labeled'
    # opt.source = '/media/personal_data/zhangq/DeepSORT/Yolov5_DeepSort_Pytorch/val_utils/data/MOT17/train/MOT17-04-FRCNN/MOT17-04-FRCNN'
    opt.yolo_model = '/mnt/ChillDisk/personal_data/zhangq/RadarRGBFusionNet2_20231128/Detection/yolov5/weights/yolov5m.pt'
    opt.deep_sort_model = 'osnet_ain_x1_0'
    opt.classes = 2
    opt.imgsz *= 2 if len(opt.imgsz) == 1 else 1  # expand

    with torch.no_grad():
        run(opt)