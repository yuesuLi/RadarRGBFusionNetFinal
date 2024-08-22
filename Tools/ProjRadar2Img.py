
import os
import numpy as np
import cv2
import torch
import openpyxl
from dataset.self_dataset2 import self_dataset2
from TiProcess.BEV_TI import read_pcd, load_json
import matplotlib.pyplot as plt


def tlwh2center(bbox_tlwh):
    if isinstance(bbox_tlwh, np.ndarray):
        bbox_center = bbox_tlwh.copy()
    elif isinstance(bbox_tlwh, torch.Tensor):
        bbox_center = bbox_tlwh.clone()
    bbox_center[:, 0] = bbox_center[:, 0] + bbox_center[:, 2] / 2.
    bbox_center[:, 1] = bbox_center[:, 1] + bbox_center[:, 3] / 2.
    return bbox_center[:, 0:2]


def Radar2img(points, OCU_json, cam_json):

    V = np.array(cam_json['intrinsic'])
    V0 = np.hstack((V, np.ones((3, 1))))        # 3*4
    RT = np.array(OCU_json['OCULiiRadar_to_LeopardCamera0_extrinsic'])  # 4*4, OCULiiRadar_to_LeopardCamera0_extrinsic, TIRadar_to_LeopardCamera0_extrinsic
    VRT = np.matmul(V0, RT)
    # length = points.shape[0]
    # xyz = np.concatenate([points[:, 0].reshape((length, 1)),
    #                       points[:, 2].reshape((length, 1)),
    #                       points[:, 1].reshape((length, 1))], axis=1).T # 3*N
    xyz = points[:, 0:3].T     # 3*N
    xyz1 = np.concatenate((xyz, np.ones((1, xyz.shape[1]))))  # 给xyz在后面叠加了一行1 (4,N)
    uvd_img = np.matmul(VRT, xyz1)  # (3,N)
    uvd_img[0, :] = uvd_img[0, :] / uvd_img[2, :]
    uvd_img[1, :] = uvd_img[1, :] / uvd_img[2, :]
    # xyz_radar2 = xyz_radar2[2, :] / xyz_radar2[2, :]

    return uvd_img


def getUVDistance(bbox_centers, radar_uvds):
    """
        bbox_centers: (N, 2): bboxes center's uv
        radar_uvds: (3, M): radar points's uvd
     """
    radar_uvs = radar_uvds[0:2, :].T  # (M, 2)
    a, b = np.asarray(bbox_centers), np.asarray(radar_uvs)
    if len(a) == 0 or len(b) == 0:
        return np.zeros((len(a), len(b)))
    a2, b2 = np.square(a).sum(axis=1), np.square(b).sum(axis=1)
    r2 = -2. * np.dot(a, b.T) + a2[:, None] + b2[None, :]
    # r2 = np.clip(r2, 0., float(np.inf))
    # make min value is 0, max value is inf
    r2 = np.sqrt(r2)    # (N, M)
    return r2

def getImgDepth(bbox_xywhs, rgb_json, OCU_json, OCU_data):

    min_num = 3
    LeopardCamera0_IntrinsicMatrix = np.array(rgb_json['intrinsic'])  # 3*3
    OCU_to_LeopardCamera0_TransformMatrix = np.array(OCU_json['OCULiiRadar_to_LeopardCamera0_extrinsic'])  # 4*4
    R = OCU_to_LeopardCamera0_TransformMatrix[0:3, 0:3] # 3*3
    T = OCU_to_LeopardCamera0_TransformMatrix[0:3, 3].reshape(3, 1) # 3*1
    radar_uvd = Radar2img(OCU_data, OCU_json, rgb_json)  # (3, M)
    radar_num = radar_uvd.shape[1]

    bbox_center = bbox_xywhs[:, 0:2]    # (N, 2)
    bbox_num = bbox_xywhs.shape[0]
    bbox_radar_distances = getUVDistance(bbox_center, radar_uvd) # (N, M)
    bbox_center_uvd = np.concatenate((bbox_center, np.zeros((bbox_num, 1))), axis=1)  # (N, 3)

    for i in range(bbox_num):
        bbox_radar_distance = bbox_radar_distances[i]
        mins_uv_indexes = np.argpartition(bbox_radar_distance, min_num)  # min_num indexes
        mins_d = radar_uvd[2, mins_uv_indexes[:min_num]]
        mean_d = np.mean(mins_d)
        bbox_center_uvd[i, 2] = mean_d

    bbox_center_uvd[:, 0] = bbox_center_uvd[:, 0] * bbox_center_uvd[:, 2]
    bbox_center_uvd[:, 1] = bbox_center_uvd[:, 1] * bbox_center_uvd[:, 2]
    bbox_center_uvd = bbox_center_uvd.T # (3, N)

    np.linalg.inv(LeopardCamera0_IntrinsicMatrix)
    XYZ_camera = np.matmul(np.linalg.inv(LeopardCamera0_IntrinsicMatrix), bbox_center_uvd)   # 3*N
    XYZ_OCU = np.matmul(np.linalg.inv(R), XYZ_camera - T)   # 3*N

    XYZ_OCU = np.concatenate([XYZ_OCU[0, :].reshape((bbox_num, 1)),
                          XYZ_OCU[2, :].reshape((bbox_num, 1)),
                          XYZ_OCU[1, :].reshape((bbox_num, 1))], axis=1) # 3*N

    print('done')
    return XYZ_OCU


def run():
    # Dataloader
    file_path = '/mnt/ChillDisk/personal_data/zhangq/RadarRGBFusionNet2_20231128/GroupPath/20230516.xlsx'
    DataPath = openpyxl.load_workbook(file_path)
    ws = DataPath.active
    groups_excel = ws['A']
    label_excel = ws['B']
    datasets_path = []
    labels_path = []
    for cell in groups_excel:
        datasets_path.append(cell.value)
    for cell in label_excel:
        labels_path.append(cell.value)
    imgsz = [640]
    imgsz *= 2 if len(imgsz) == 1 else 1  # expand

    for i in range(len(datasets_path)):
        data_base_path = '/mnt/ourDataset_v2/ourDataset_v2_label'
        source = os.path.join(data_base_path, datasets_path[i])

        dataset = self_dataset2(source, img_size=imgsz)


    # OCU_data: x y z doppler snr (N, 5)
        for frame_idx, (rgb_img0, img, rgb_json, OCU_data, OCU_json, lidar_json, curr_label, Lidar_points) in enumerate(dataset):
            print('frame_num:', frame_idx)
            radar_uvd = Radar2img(OCU_data, OCU_json, rgb_json)  # 3*N
            radar_num = radar_uvd.shape[1]
            bbox_xywhs = np.array([[500, 200, 1], [400, 300, 2], [600, 100, 3], [550, 250, 4]])
            bbox_proj_proj = getImgDepth(bbox_xywhs, rgb_json, OCU_json, OCU_data)  # 3*N
            for i in range(radar_num):
                cv2.circle(rgb_img0, (int(radar_uvd[0, i]), int(radar_uvd[1, i])), radius=5, color=(0, 0, 255),
                           thickness=-1)

            # bbox_tlwh = 0   # (N, 4)
            # bbox_num = bbox_tlwh.shape[0]
            # bbox_center = tlwh2center(bbox_tlwh)    # (N, 2)
            # center_uvd = np.concatenate([bbox_center, np.zeros((bbox_num, 1))])  #  (N, 3)
            # for i in range(bbox_num):
            #     for j in range(radar_num):







            # if frame_idx == 138:
            #     cv2.imshow('test', rgb_img0)
            #     if cv2.waitKey(0) & 0xFF == 27:
            #         break
            #     print()
            # cv2.imshow('test', rgb_img0)
            # cv2.waitKey(100)
            # if cv2.waitKey(0) & 0xFF == 27:
            #     break

    print('done')


if __name__ == '__main__':
    run()